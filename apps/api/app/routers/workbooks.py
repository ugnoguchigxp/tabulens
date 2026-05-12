import os
import uuid
import pandas as pd
from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from openpyxl import load_workbook as load_openpyxl_workbook
from app.models.schemas import (
    WorkbookUploadResponse,
    SheetInfo,
    ColumnInfo,
    SheetRowsResponse,
    SheetProfileResponse,
    WorkbookFormulaMetadataResponse,
)
from app.core.paths import UPLOAD_DIR
from app.services.workbook_formula_store import (
    load_workbook_formula_metadata,
    save_workbook_formula_metadata,
)

router = APIRouter()

@router.post("/upload", response_model=WorkbookUploadResponse)
async def upload_workbook(file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in [".xlsx", ".csv"]:
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are supported")
    
    workbook_id = str(uuid.uuid4())
    file_path = UPLOAD_DIR / f"{workbook_id}{ext}"
    
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
    
    try:
        sheets_info = []
        
        if ext == ".xlsx":
            excel_file = pd.ExcelFile(file_path)
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheets_info.append(_process_dataframe(df, sheet_name))
            formula_metadata = _extract_formula_metadata(file_path, workbook_id)
            save_workbook_formula_metadata(workbook_id, formula_metadata)
        else:
            # Handle CSV as a single sheet
            df = pd.read_csv(file_path)
            sheets_info.append(_process_dataframe(df, "CSV Data"))
            save_workbook_formula_metadata(workbook_id, {"workbook_id": workbook_id, "sheets": []})
            
        return WorkbookUploadResponse(
            workbook_id=workbook_id,
            sheets=sheets_info
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@router.get("/{workbook_id}", response_model=WorkbookUploadResponse)
async def get_workbook(workbook_id: str):
    for ext in [".xlsx", ".csv"]:
        file_path = UPLOAD_DIR / f"{workbook_id}{ext}"
        if file_path.exists():
            try:
                return _load_workbook(file_path, workbook_id)
            except Exception as exc:
                raise HTTPException(status_code=500, detail=f"Error reading workbook: {exc}") from exc
    raise HTTPException(status_code=404, detail="Workbook not found")


@router.get("/{workbook_id}/sheets/{sheet_name}/preview")
async def preview_sheet(workbook_id: str, sheet_name: str):
    for ext in [".xlsx", ".csv"]:
        file_path = UPLOAD_DIR / f"{workbook_id}{ext}"
        if not file_path.exists():
            continue
        try:
            if ext == ".xlsx":
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                if sheet_name not in {"CSV Data", "Sheet1"}:
                    continue
                df = pd.read_csv(file_path)
            preview_df = df.head(10).fillna("")
            return {
                "workbook_id": workbook_id,
                "sheet_name": sheet_name,
                "preview_rows": preview_df.to_dict(orient="records"),
            }
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Error creating preview: {exc}") from exc
    raise HTTPException(status_code=404, detail="Workbook not found")


@router.get("/{workbook_id}/formulas", response_model=WorkbookFormulaMetadataResponse)
async def workbook_formulas(workbook_id: str):
    metadata = load_workbook_formula_metadata(workbook_id)
    if metadata is not None:
        return WorkbookFormulaMetadataResponse.model_validate(metadata)

    file_path = _resolve_workbook_path(workbook_id)
    if file_path is None:
        raise HTTPException(status_code=404, detail="Workbook not found")

    if str(file_path).endswith(".xlsx"):
        metadata = _extract_formula_metadata(file_path, workbook_id)
    else:
        metadata = {"workbook_id": workbook_id, "sheets": []}
    save_workbook_formula_metadata(workbook_id, metadata)
    return WorkbookFormulaMetadataResponse.model_validate(metadata)


@router.get("/{workbook_id}/sheets/{sheet_name}/rows", response_model=SheetRowsResponse)
async def sheet_rows(
    workbook_id: str,
    sheet_name: str,
    offset: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
):
    df = _load_sheet_or_404(workbook_id, sheet_name)
    row_count = int(len(df))
    rows = df.iloc[offset:offset + limit].fillna("").to_dict(orient="records")
    return SheetRowsResponse(
        workbook_id=workbook_id,
        sheet_name=sheet_name,
        offset=offset,
        limit=limit,
        row_count=row_count,
        rows=rows,
    )


@router.get("/{workbook_id}/sheets/{sheet_name}/profile", response_model=SheetProfileResponse)
async def sheet_profile(workbook_id: str, sheet_name: str):
    df = _load_sheet_or_404(workbook_id, sheet_name)
    row_count = int(len(df))
    column_count = int(len(df.columns))
    missing_rate_overall = float(df.isna().mean().mean()) if row_count and column_count else 0.0
    return SheetProfileResponse(
        workbook_id=workbook_id,
        sheet_name=sheet_name,
        row_count=row_count,
        column_count=column_count,
        missing_rate_overall=missing_rate_overall,
    )

def _process_dataframe(df: pd.DataFrame, sheet_name: str) -> SheetInfo:
    columns = []
    for col in df.columns:
        inferred_type = str(df[col].dtype)
        missing_count = int(df[col].isna().sum())
        columns.append(ColumnInfo(
            name=str(col),
            inferred_type=inferred_type,
            missing_count=missing_count
        ))

    display_df = df.fillna("")
    preview_rows = display_df.head(10).to_dict(orient="records")
    
    return SheetInfo(
        name=sheet_name,
        row_count=len(df),
        columns=columns,
        preview_rows=preview_rows
    )


def _load_workbook(file_path, workbook_id: str) -> WorkbookUploadResponse:
    sheets_info = []
    if str(file_path).endswith(".xlsx"):
        excel_file = pd.ExcelFile(file_path)
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            sheets_info.append(_process_dataframe(df, sheet_name))
    else:
        df = pd.read_csv(file_path)
        sheets_info.append(_process_dataframe(df, "CSV Data"))
    return WorkbookUploadResponse(workbook_id=workbook_id, sheets=sheets_info)


def _resolve_workbook_path(workbook_id: str):
    for ext in [".xlsx", ".csv"]:
        file_path = UPLOAD_DIR / f"{workbook_id}{ext}"
        if file_path.exists():
            return file_path
    return None


def _extract_formula_metadata(file_path, workbook_id: str) -> dict:
    workbook = load_openpyxl_workbook(file_path, data_only=False)
    workbook_cached = load_openpyxl_workbook(file_path, data_only=True)
    sheets: list[dict] = []

    for worksheet in workbook.worksheets:
        cached_sheet = workbook_cached[worksheet.title]
        cells: list[dict] = []
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if cell.data_type != "f" and not (isinstance(value, str) and value.startswith("=")):
                    continue
                formula_text = value if isinstance(value, str) and value.startswith("=") else f"={value}"
                cached_value = cached_sheet[cell.coordinate].value
                cells.append({
                    "address": cell.coordinate,
                    "formula": formula_text,
                    "cached_value": cached_value,
                })
        sheets.append({"name": worksheet.title, "cells": cells})

    return {"workbook_id": workbook_id, "sheets": sheets}


def _load_sheet_or_404(workbook_id: str, sheet_name: str) -> pd.DataFrame:
    for ext in [".xlsx", ".csv"]:
        file_path = UPLOAD_DIR / f"{workbook_id}{ext}"
        if not file_path.exists():
            continue
        try:
            if ext == ".xlsx":
                excel_file = pd.ExcelFile(file_path)
                if sheet_name not in excel_file.sheet_names:
                    raise HTTPException(status_code=404, detail="Sheet not found")
                return pd.read_excel(file_path, sheet_name=sheet_name)
            if sheet_name in {"CSV Data", "Sheet1"}:
                return pd.read_csv(file_path)
        except Exception as exc:
            if isinstance(exc, HTTPException):
                raise
            raise HTTPException(status_code=500, detail=f"Error reading sheet: {exc}") from exc
    raise HTTPException(status_code=404, detail="Workbook not found")
